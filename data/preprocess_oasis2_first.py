"""
preprocess_oasis2.py
OASIS-2 Preprocessing Pipeline — CS4100 Group Project
Ilay Zubkov, Jacob Shechter, Francesca Caldarella

Produces fixed-size, normalised 3D volumes and a train/val/test split manifest
from the raw OASIS-2 MRI data. Run once — outputs are saved and reused.

NOTE: Additional preprocessing steps (e.g. non-brain tissue removal, intensity
artifact correction) are standard practice and will be added in a future
iteration once the baseline model is established.
"""

import json
import random
from pathlib import Path
from collections import Counter

import numpy as np
import nibabel as nib
import openpyxl
from scipy.ndimage import zoom

BASE        = Path(__file__).parent
PART1       = BASE / "OAS2_RAW_PART1"
PART2       = BASE / "OAS2_RAW_PART2"
DEMO_FILE   = BASE / "oasis_longitudinal_demographics-8d83e569fa2e2d30.xlsx"
OUT_DIR     = BASE / "data" / "processed"
SPLITS_FILE = BASE / "data" / "splits.json"

TARGET_SHAPE = (96, 96, 96)
CDR_TO_CLASS = {0: 0, 0.5: 1, 1: 2, 2: 2}
RANDOM_SEED  = 42


# Step 1: Load CDR labels and metadata from Excel

def load_demographics(path):
    wb      = openpyxl.load_workbook(path)
    ws      = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    records = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        r      = dict(zip(headers, row))
        mri_id = r["MRI ID"]
        cdr    = r["CDR"]
        if mri_id is None or cdr is None:
            continue
        cdr = float(cdr)
        records[mri_id] = {
            "subject_id": r["Subject ID"],
            "cdr":        cdr,
            "label":      CDR_TO_CLASS[cdr],
            "group":      r["Group"],
            "age":        r["Age"],
            "sex":        r["M/F"],
            "mmse":       r["MMSE"],
            "etiv":       r["eTIV"],
            "nwbv":       r["nWBV"],
        }
    return records


# Step 2: Collect session folder paths from PART1 and PART2

def collect_sessions(part_dirs):
    sessions = {}
    for part in part_dirs:
        for d in sorted(part.iterdir()):
            if d.is_dir():
                sessions[d.name] = d
    return sessions


# Steps 3–4: Load all mpr-N acquisitions per session and average them

def load_and_average(session_dir):
    hdr_files = sorted((session_dir / "RAW").glob("mpr-*.nifti.hdr"))
    if not hdr_files:
        raise FileNotFoundError(f"No acquisitions in {session_dir / 'RAW'}")
    volumes = []
    for hdr in hdr_files:
        data = nib.load(str(hdr)).get_fdata(dtype=np.float32)
        if data.ndim == 4:
            data = data.squeeze(-1)
        volumes.append(data)
    return np.mean(volumes, axis=0).astype(np.float32)


# Step 5: Resample to TARGET_SHAPE using trilinear interpolation

def resample(volume, target):
    factors = tuple(t / s for t, s in zip(target, volume.shape))
    return zoom(volume, factors, order=1).astype(np.float32)


# Step 6: Z-score normalise per volume, clip outliers to [-5, 5]

def zscore_normalise(volume):
    mu, std = volume.mean(), volume.std()
    if std < 1e-6:
        return np.zeros_like(volume)
    return np.clip((volume - mu) / std, -5.0, 5.0).astype(np.float32)


# Steps 7–8: Map CDR to 3-class labels, split by subject (70/15/15)

def subject_level_split(records):
    subjects = sorted(set(r["subject_id"] for r in records))
    random.seed(RANDOM_SEED)
    random.shuffle(subjects)
    n      = len(subjects)
    n_val  = max(1, round(n * 0.15))
    n_test = max(1, round(n * 0.15))
    test_s  = set(subjects[:n_test])
    val_s   = set(subjects[n_test : n_test + n_val])
    train_s = set(subjects[n_test + n_val :])
    splits  = {"train": [], "val": [], "test": []}
    for r in records:
        if r["subject_id"] in train_s:
            splits["train"].append(r)
        elif r["subject_id"] in val_s:
            splits["val"].append(r)
        else:
            splits["test"].append(r)
    return splits


# Step 9: Run pipeline, save .npy volumes and splits.json

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    demo     = load_demographics(DEMO_FILE)
    sessions = collect_sessions([PART1, PART2])
    valid    = {mid: sessions[mid] for mid in sessions if mid in demo}

    print(f"Sessions found: {len(valid)} / {len(demo)}")

    records, errors = [], []
    for i, (mri_id, session_dir) in enumerate(sorted(valid.items())):
        out_path = OUT_DIR / f"{mri_id}.npy"
        if out_path.exists():
            records.append({**demo[mri_id], "mri_id": mri_id, "path": str(out_path)})
            continue
        try:
            vol = load_and_average(session_dir)
            vol = resample(vol, TARGET_SHAPE)
            vol = zscore_normalise(vol)
            np.save(out_path, vol)
            records.append({**demo[mri_id], "mri_id": mri_id, "path": str(out_path)})
            if (i + 1) % 50 == 0 or (i + 1) == len(valid):
                print(f"  [{i+1}/{len(valid)}] {mri_id}")
        except Exception as e:
            errors.append((mri_id, str(e)))
            print(f"  [ERROR] {mri_id}: {e}")

    print(f"Done: {len(records)} processed, {len(errors)} errors")

    splits = subject_level_split(records)
    for name, recs in splits.items():
        lbl = Counter(r["label"] for r in recs)
        print(f"  {name:5s}: {len(recs)} sessions | "
              f"class 0={lbl[0]}  class 1={lbl[1]}  class 2={lbl[2]}")

    SPLITS_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(SPLITS_FILE, "w") as f:
        json.dump(splits, f, indent=2)
    print(f"Splits saved -> {SPLITS_FILE}")


if __name__ == "__main__":
    main()
