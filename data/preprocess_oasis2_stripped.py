"""
preprocess_oasis2_stripped.py
OASIS-2 Preprocessing Pipeline with Skull Stripping
CS4100 Group Project — Ilay Zubkov, Jacob Shechter, Francesca Caldarella

Identical to preprocess_oasis2.py with one key addition: skull stripping.
Brain extraction is applied on raw intensities (before normalization) using
antspynet.brain_extraction (Tustison et al. 2021), a pretrained deep learning
model cited as a preprocessing tool, not our own implementation.

WHY ORDER MATTERS:
  Skull stripping must run on raw intensities — the model was trained on
  raw MRI scanner values and cannot identify brain tissue in normalized data.
  Normalization then runs only on brain voxels, so background zeros don't
  skew the mean/std calculation.

Pipeline:
  1. Load CDR labels and demographics
  2. Collect session folder paths
  3. Load and average mpr-N acquisitions per session
  4. Resample to 96x96x96 (on raw intensities)
  5. Skull strip — extract brain mask from raw intensities
  6. Apply mask — zero out non-brain voxels
  7. Z-score normalize — computed on brain voxels only
  8. CDR label mapping + subject-level split
  9. Save volumes to data/processed_stripped/
 10. Save split manifest to data/splits.json (same split as before)

Install dependencies (Colab/Kaggle):
  !pip install antspyx antspynet nibabel openpyxl scipy -q
"""

import json
import random
from pathlib import Path
from collections import Counter

import numpy as np
import nibabel as nib
import openpyxl
from scipy.ndimage import zoom
import ants
from antspynet.utilities import brain_extraction

# CONFIGURATION
# Override these paths when running on Colab or Kaggle

# This file lives in data/, so .parent.parent resolves to the project root
# where the raw OASIS-2 archives are extracted.
BASE        = Path(__file__).parent.parent
PART1       = BASE / "OAS2_RAW_PART1"
PART2       = BASE / "OAS2_RAW_PART2"
DEMO_FILE   = BASE / "oasis_longitudinal_demographics-8d83e569fa2e2d30.xlsx"
OUT_DIR     = BASE / "data" / "processed_stripped"
SPLITS_FILE = BASE / "data" / "splits.json"

TARGET_SHAPE = (96, 96, 96)
CDR_TO_CLASS = {0: 0, 0.5: 1, 1: 2, 2: 2}
RANDOM_SEED  = 42   # same seed as original — produces identical split


# STEP 1: LOAD DEMOGRAPHICS

def load_demographics(path: Path) -> dict:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

    records = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        r      = dict(zip(headers, row))
        mri_id = r["MRI ID"]
        cdr    = r["CDR"]
        if mri_id is None or cdr is None:
            continue
        cdr = float(cdr)
        records[mri_id] = {
            "subject_id": r["Subject ID"],
            "cdr":        cdr,
            "label":      CDR_TO_CLASS[cdr],
            "group":      r["Group"],
            "age":        r["Age"],
            "sex":        r["M/F"],
            "mmse":       r["MMSE"],
            "etiv":       r["eTIV"],
            "nwbv":       r["nWBV"],
        }
    return records


# STEP 2: COLLECT SESSION PATHS

def collect_sessions(part_dirs: list) -> dict:
    sessions = {}
    for part in part_dirs:
        for session_dir in sorted(part.iterdir()):
            if session_dir.is_dir():
                sessions[session_dir.name] = session_dir
    return sessions


# STEP 3+4: LOAD AND AVERAGE ACQUISITIONS

def load_and_average(session_dir: Path) -> np.ndarray:
    """Returns mean of all mpr-N acquisitions, shape (256, 256, 128), raw intensities."""
    raw_dir   = session_dir / "RAW"
    hdr_files = sorted(raw_dir.glob("mpr-*.nifti.hdr"))
    if not hdr_files:
        raise FileNotFoundError(f"No acquisition files found in {raw_dir}")
    volumes = []
    for hdr in hdr_files:
        img  = nib.load(str(hdr))
        data = img.get_fdata(dtype=np.float32)
        if data.ndim == 4:
            data = data.squeeze(-1)
        volumes.append(data)
    return np.mean(volumes, axis=0).astype(np.float32)


# STEP 5: RESAMPLE

def resample(volume: np.ndarray, target: tuple) -> np.ndarray:
    """Trilinear resampling to target shape."""
    factors = tuple(t / s for t, s in zip(target, volume.shape))
    return zoom(volume, factors, order=1).astype(np.float32)


# STEP 6: SKULL STRIP

def skull_strip(volume: np.ndarray) -> np.ndarray:
    """
    Runs antspynet brain extraction on a raw-intensity volume.
    Returns a binary mask (1 = brain, 0 = background).

    Must be called on raw intensities — antspynet's model was trained on
    raw MRI scanner values and cannot identify brain tissue in normalized data.
    """
    ants_img    = ants.from_numpy(volume, spacing=(1.0, 1.0, 1.0))
    prob_mask   = brain_extraction(ants_img, modality="t1")
    binary_mask = (prob_mask.numpy() > 0.5).astype(np.float32)
    return binary_mask


# STEP 7: NORMALIZE BRAIN VOXELS ONLY

def zscore_normalise_brain(volume: np.ndarray, mask: np.ndarray) -> np.ndarray:
    """
    Z-score normalization computed only over brain voxels (mask == 1).
    Background voxels stay at exactly 0.

    Computing stats over the whole volume (including background zeros) would
    pull the mean toward zero and underestimate the true brain intensity spread.
    """
    brain_voxels = volume[mask == 1]

    if brain_voxels.size == 0:
        return np.zeros_like(volume)

    mean = brain_voxels.mean()
    std  = brain_voxels.std()

    if std < 1e-6:
        return np.zeros_like(volume)

    normalized          = np.zeros_like(volume)
    normalized[mask==1] = (brain_voxels - mean) / std
    return np.clip(normalized, -5.0, 5.0).astype(np.float32)


# STEP 8: SUBJECT-LEVEL SPLIT

def subject_level_split(records: list, seed: int = RANDOM_SEED):
    """Same 70/15/15 subject-level split as the original preprocessing."""
    subject_ids = sorted(set(r["subject_id"] for r in records))
    random.seed(seed)
    random.shuffle(subject_ids)

    n      = len(subject_ids)
    n_val  = max(1, round(n * 0.15))
    n_test = max(1, round(n * 0.15))

    test_subs  = set(subject_ids[:n_test])
    val_subs   = set(subject_ids[n_test : n_test + n_val])
    train_subs = set(subject_ids[n_test + n_val :])

    splits = {"train": [], "val": [], "test": []}
    for r in records:
        if r["subject_id"] in train_subs:
            splits["train"].append(r)
        elif r["subject_id"] in val_subs:
            splits["val"].append(r)
        else:
            splits["test"].append(r)
    return splits


# MAIN

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Step 1: Loading demographics")
    demo = load_demographics(DEMO_FILE)
    print(f"  Sessions in spreadsheet: {len(demo)}")

    print("\nStep 2: Collecting session paths")
    sessions = collect_sessions([PART1, PART2])
    valid    = {mid: sessions[mid] for mid in sessions if mid in demo}
    print(f"  Sessions with labels: {len(valid)}")

    print("\nSteps 3-7: Averaging, resampling, skull stripping, normalizing")
    records = []
    errors  = []

    for i, (mri_id, session_dir) in enumerate(sorted(valid.items())):
        out_path = OUT_DIR / f"{mri_id}.npy"

        if out_path.exists():
            records.append({**demo[mri_id], "mri_id": mri_id,
                            "path": str(out_path)})
            print(f"  [{i+1:>3}/{len(valid)}] skip (done): {mri_id}")
            continue

        try:
            vol  = load_and_average(session_dir)   # raw intensities (256,256,128)
            vol  = resample(vol, TARGET_SHAPE)      # raw intensities (96,96,96)
            mask = skull_strip(vol)                 # binary brain mask
            vol  = vol * mask                       # zero out skull
            vol  = zscore_normalise_brain(vol, mask)# normalize brain voxels only

            np.save(out_path, vol)
            records.append({**demo[mri_id], "mri_id": mri_id,
                            "path": str(out_path)})

            brain_pct = 100 * mask.mean()
            print(f"  [{i+1:>3}/{len(valid)}] {mri_id}  brain={brain_pct:.1f}%")

        except Exception as e:
            errors.append((mri_id, str(e)))
            print(f"  [ERROR] {mri_id}: {e}")

    print(f"\n  Processed: {len(records)}  Errors: {len(errors)}")

    print("\nSteps 8-10: Split and save")
    splits = subject_level_split(records)
    for split_name, split_records in splits.items():
        lbl = Counter(r["label"] for r in split_records)
        print(f"  {split_name:5s}: {len(split_records):3d} sessions | "
              f"c0={lbl[0]}  c1={lbl[1]}  c2={lbl[2]}")

    SPLITS_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(SPLITS_FILE, "w") as f:
        json.dump(splits, f, indent=2)
    print(f"\n  Splits saved -> {SPLITS_FILE}")
    print(f"  Volumes saved -> {OUT_DIR}")
    print("\nDone.")


if __name__ == "__main__":
    main()
